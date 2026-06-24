import streamlit as st
import pandas as pd
import re
from io import BytesIO
from datetime import datetime

from openpyxl.styles import (
    Alignment,
    Border,
    Side,
    PatternFill,
    Font
)

from openpyxl.utils import get_column_letter


# =========================
# 頁面設定
# =========================

st.set_page_config(
    page_title="櫃位轉檔工具",
    page_icon="📊",
    layout="centered"
)

st.title("📊 櫃位轉檔工具")
st.write("請上傳訂單報表（Excel 或 CSV）")


# =========================
# 轉檔主函式
# =========================

def convert_report(uploaded_file):

    file_ext = uploaded_file.name.lower().split(".")[-1]

    if file_ext == "csv":
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)

    if "POS商店名稱" not in df.columns:
        raise Exception("找不到 POS商店名稱 欄位")

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        # 保留原始訂單明細
        df.to_excel(
            writer,
            sheet_name="訂單明細",
            index=False
        )

        grouped = df.groupby("POS商店名稱")

        apparel_sizes = [
            "XS",
            "S",
            "M",
            "L",
            "XL",
            "XXL"
        ]

        total_store = len(grouped)

        progress_bar = st.progress(0)

        current_store = 0

        for store_name, group in grouped:

            current_store += 1

            progress_bar.progress(
                current_store / total_store
            )

            if pd.isna(store_name):
                continue

            if str(store_name).strip() == "":
                continue

            store_data = {}

            for _, row in group.iterrows():

                vendor_code = str(
                    row.get("廠商編號", "")
                ).strip()

                qty = row.get("數量", 1)

                try:
                    qty = int(qty)
                except:
                    qty = 1

                if (
                    pd.isna(row.get("廠商編號"))
                    or vendor_code == ""
                    or vendor_code == "nan"
                ):
                    continue

                if "_" in vendor_code:

                    model_color, size_val = (
                        vendor_code.rsplit("_", 1)
                    )

                    size_val = size_val.strip()

                else:

                    model_color = vendor_code
                    size_val = ""

                if model_color not in store_data:

                    store_data[model_color] = {
                        "sizes": [""] * 17,
                        "unmapped_qty": 0
                    }

                col_idx = None

                if size_val in apparel_sizes:

                    col_idx = (
                        apparel_sizes.index(size_val)
                        + 3
                    )

                else:

                    model_part = (
                        model_color.rsplit("-", 1)[0]
                        if "-" in model_color
                        else model_color
                    )

                    match = re.search(
                        r"([a-zA-Z]+)$",
                        model_part
                    )

                    suffix = (
                        match.group(1).upper()
                        if match
                        else ""
                    )

                    try:

                        size_int = int(size_val)

                        if (
                            suffix == "B"
                            and 24 <= size_int <= 40
                        ):
                            col_idx = size_int - 24

                        elif (
                            suffix in [
                                "M",
                                "W",
                                "BW",
                                ""
                            ]
                            and 34 <= size_int <= 48
                        ):
                            col_idx = size_int - 34

                    except ValueError:
                        pass

                if (
                    col_idx is not None
                    and 0 <= col_idx < 17
                ):

                    current_qty = (
                        store_data[model_color]["sizes"][col_idx]
                    )

                    store_data[model_color]["sizes"][col_idx] = (
                        qty
                        if current_qty == ""
                        else current_qty + qty
                    )

                else:

                    store_data[model_color][
                        "unmapped_qty"
                    ] += qty

            rows_list = []

            for mc, item_info in store_data.items():

                sizes_list = item_info["sizes"]

                row_total = (
                    sum(
                        [
                            val
                            for val in sizes_list
                            if isinstance(val, int)
                        ]
                    )
                    + item_info["unmapped_qty"]
                )

                rows_list.append(
                    [mc]
                    + sizes_list
                    + [row_total]
                )

            if not rows_list:
                continue

            rows_list.sort(
                key=lambda x: str(x[0])
            )

            df_output = pd.DataFrame(
                rows_list
            )

            sheet_name = str(store_name)[:31]

            df_output.to_excel(
                writer,
                sheet_name=sheet_name,
                startrow=3,
                index=False,
                header=False,
            )

            worksheet = writer.sheets[
                sheet_name
            ]

            # ======================
            # 客製表頭
            # ======================

            worksheet.cell(
                row=1,
                column=1,
                value=sheet_name
            )

            worksheet.merge_cells(
                "A1:A2"
            )

            for i in range(15):

                worksheet.cell(
                    row=1,
                    column=i + 2,
                    value=str(34 + i)
                )

            for i in range(17):

                worksheet.cell(
                    row=2,
                    column=i + 2,
                    value=str(24 + i)
                )

            worksheet.cell(
                row=3,
                column=1,
                value="商品型號-色號"
            )

            for i, s in enumerate(
                apparel_sizes
            ):

                worksheet.cell(
                    row=3,
                    column=i + 5,
                    value=s
                )

            worksheet.cell(
                row=3,
                column=19,
                value="合計"
            )

            # ======================
            # 格式設定
            # ======================

            worksheet.freeze_panes = "A4"

            worksheet.column_dimensions[
                "A"
            ].width = 14.5

            for col_idx in range(
                2,
                20
            ):

                col_letter = (
                    get_column_letter(
                        col_idx
                    )
                )

                worksheet.column_dimensions[
                    col_letter
                ].width = 5.0

            align_center = Alignment(
                horizontal="center",
                vertical="center"
            )

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            green_fill = PatternFill(
                start_color="CCFFCC",
                end_color="CCFFCC",
                fill_type="solid"
            )

            bold_font = Font(
                bold=True
            )

            max_row = worksheet.max_row

            for row_idx in range(
                1,
                max_row + 1
            ):

                for col_idx in range(
                    1,
                    20
                ):

                    cell = worksheet.cell(
                        row=row_idx,
                        column=col_idx
                    )

                    cell.alignment = (
                        align_center
                    )

                    if row_idx >= 4:

                        cell.border = (
                            thin_border
                        )

                    elif (
                        cell.value
                        is not None
                    ):

                        if (
                            str(cell.value)
                            .strip()
                            != ""
                        ):

                            cell.border = (
                                thin_border
                            )

                    if row_idx <= 3:

                        cell.fill = (
                            green_fill
                        )

                        cell.font = (
                            bold_font
                        )

    output.seek(0)

    return output, df
# =========================
# Streamlit UI
# =========================

uploaded_file = st.file_uploader(
    "選擇檔案",
    type=["xlsx", "csv"]
)

if uploaded_file:

    st.success("檔案上傳成功")

    try:

        if uploaded_file.name.lower().endswith(
            ".csv"
        ):
            preview_df = pd.read_csv(
                uploaded_file
            )
        else:
            preview_df = pd.read_excel(
                uploaded_file
            )

        st.info(
            f"檔案名稱：{uploaded_file.name}"
        )

        st.info(
            f"資料筆數：{len(preview_df):,}"
        )

        if (
            "POS商店名稱"
            in preview_df.columns
        ):

            store_count = (
                preview_df[
                    "POS商店名稱"
                ]
                .dropna()
                .nunique()
            )

            st.info(
                f"專櫃數量：{store_count}"
            )

        col1, col2 = st.columns(2)

        with col1:

            st.dataframe(
                preview_df.head(10),
                use_container_width=True
            )

        with col2:

            st.write("欄位列表")

            st.dataframe(
                pd.DataFrame(
                    preview_df.columns,
                    columns=["欄位名稱"]
                ),
                use_container_width=True
            )

        if st.button(
            "🚀 開始轉換",
            type="primary",
            use_container_width=True
        ):

            with st.spinner(
                "轉換中，請稍候..."
            ):

                uploaded_file.seek(0)

                output, result_df = (
                    convert_report(
                        uploaded_file
                    )
                )

            st.success(
                "✅ 轉換完成"
            )

            file_name = (
                f"{datetime.now().strftime('%Y%m%d')}"
                "櫃位轉檔.xlsx"
            )

            st.download_button(
                label="📥 下載櫃位轉檔",
                data=output,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    except Exception as e:

        st.error(
            f"錯誤：{str(e)}"
        )